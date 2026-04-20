"""Importador de clientes desde Excel (.xlsx).

Se usa desde:
- La acción 'Importar Excel' en el admin (ClienteAdmin.importar_excel_view).
- El management command `python manage.py importar_excel archivo.xlsx`.
"""
import re
import unicodedata
import openpyxl

from .models import Region, Comuna


# ───────────────────────────────────────────────────────────────────────────
#  NORMALIZACIÓN DE COMUNAS
# ───────────────────────────────────────────────────────────────────────────
def _normalizar(s):
    if not s:
        return ''
    s = str(s).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFKD', s) if not unicodedata.combining(c))
    return s.rstrip('.')


# Mapa de entradas sucias del Excel real → nombre canónico de comuna
COMUNAS_ALIAS = {
    'anfofagasta':           'Antofagasta',
    'antofagasta':           'Antofagasta',
    'incdependencia':        'Independencia',
    'independencia':         'Independencia',
    'nunoa':                 'Ñuñoa',
    'penalolen':             'Peñalolén',
    'maipu':                 'Maipú',
    'la serena':             'La Serena',
    'quilpue':               'Quilpué',
    'quilpua':               'Quilpué',
    'villa alemana':         'Villa Alemana',
    'villa alemana/quilpua': 'Villa Alemana',
    'iquique':               'Iquique',
    'calama':                'Calama',
    'santiago':              'Santiago',
    'providencia':           'Providencia',
    'las condes':            'Las Condes',
    'la reina':              'La Reina',
    'la florida':            'La Florida',
    'macul':                 'Macul',
    'huechuraba':            'Huechuraba',
    'recoleta':              'Recoleta',
    'san bernardo':          'San Bernardo',
    'la cisterna':           'La Cisterna',
    'padre hurtado':         'Padre Hurtado',
    'estacion central':      'Estación Central',
    'quilicura':             'Quilicura',
    'cerrillos':             'Cerrillos',
    'talagante':             'Talagante',
    'linderos':              'Buin',
    'san antonio':           'San Antonio',
    'rancagua':              'Rancagua',
    'pichilemu':             'Pichilemu',
    'pichidehua':            'Pichidegua',
    'pichidegua':            'Pichidegua',
    'molina':                'Molina',
    'molina, r maule':       'Molina',
    'concepcion':            'Concepción',
    'temuco':                'Temuco',
    'valdivia':              'Valdivia',
}


def buscar_comuna(raw):
    """Intenta encontrar una Comuna existente en la BD desde un texto 'sucio'."""
    if not raw:
        return None
    n = _normalizar(raw)
    canonico = COMUNAS_ALIAS.get(n)
    if canonico:
        return Comuna.objects.filter(nombre__iexact=canonico).first()
    # búsqueda difusa (comienza con…)
    posibles = Comuna.objects.filter(nombre__istartswith=raw.strip()[:6])
    if posibles.count() == 1:
        return posibles.first()
    # búsqueda exacta case-insensitive
    return Comuna.objects.filter(nombre__iexact=raw.strip()).first()


# ───────────────────────────────────────────────────────────────────────────
#  TELÉFONO
# ───────────────────────────────────────────────────────────────────────────
def formatear_telefono(raw):
    if not raw:
        return ''
    d = re.sub(r'\D', '', str(raw))
    if d.startswith('56'):
        d = d[2:]
    if len(d) == 9 and d.startswith('9'):
        return f'+56 {d[0]} {d[1:5]} {d[5:]}'
    if len(d) == 8:
        return f'+56 9 {d[:4]} {d[4:]}'
    return str(raw).strip()


# ───────────────────────────────────────────────────────────────────────────
#  PROCESADOR PRINCIPAL
# ───────────────────────────────────────────────────────────────────────────
def procesar_excel(archivo):
    """Lee el .xlsx y devuelve (lista_preview, lista_errores).

    preview: lista de dicts por fila con valido, _warn, y los campos.
    errores: lista de strings de errores globales (columna faltante, etc.)
    """
    errores = []
    wb = openpyxl.load_workbook(archivo, data_only=True)
    # primera hoja con datos
    ws = None
    for sn in wb.sheetnames:
        h = wb[sn]
        if h.max_row and h.max_row > 1:
            ws = h
            break
    if not ws:
        errores.append('El archivo no tiene hojas con datos.')
        return [], errores

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        errores.append('La hoja está vacía.')
        return [], errores

    headers = [(_normalizar(c) if c else '') for c in rows[0]]

    # Mapear columnas
    def find_col(*keys):
        for k in keys:
            for i, h in enumerate(headers):
                if k in h:
                    return i
        return -1

    idx = {
        'razon':     find_col('razon', 'razón', 'nombre'),
        'direccion': find_col('direccion', 'direcci'),
        'comuna':    find_col('comuna'),
        'contacto':  find_col('contacto'),
        'telefono':  find_col('fono', 'tel'),
        'equipos':   find_col('equipo'),
        'eco':       find_col('eco'),
    }

    if idx['razon'] < 0:
        errores.append("No se encuentra la columna 'Razón social' (ni 'nombre').")
        return [], errores
    if idx['comuna'] < 0:
        errores.append("No se encuentra la columna 'Comuna'.")
        return [], errores

    preview = []
    for r in rows[1:]:
        if not r or not any(r):
            continue
        razon  = str(r[idx['razon']] or '').strip()
        if not razon:
            continue
        direccion = str(r[idx['direccion']] or '').strip() if idx['direccion'] >= 0 else ''
        comuna_raw = str(r[idx['comuna']] or '').strip() if idx['comuna'] >= 0 else ''
        contacto = str(r[idx['contacto']] or '').strip() if idx['contacto'] >= 0 else ''
        tel      = str(r[idx['telefono']] or '').strip() if idx['telefono'] >= 0 else ''
        equipos  = str(r[idx['equipos']] or '').strip() if idx['equipos'] >= 0 else ''
        eco_raw  = str(r[idx['eco']] or '').strip().lower() if idx['eco'] >= 0 else ''

        comuna_obj = buscar_comuna(comuna_raw)

        row = {
            'razon_social': razon,
            'direccion': direccion,
            'comuna_raw': comuna_raw,
            'comuna': comuna_obj.nombre if comuna_obj else comuna_raw,
            'region': comuna_obj.region.nombre if comuna_obj else 'Desconocida',
            'contacto': contacto,
            'telefono': formatear_telefono(tel),
            'equipos': equipos,
            'eco_friendly': eco_raw in ('si', 'sí'),
            '_comuna_obj': comuna_obj,
            'valido': True,
            '_warn': '',
        }
        if not comuna_obj:
            row['valido'] = False
            row['_warn']   = f"Comuna '{comuna_raw}' no encontrada"
        preview.append(row)

    return preview, errores
